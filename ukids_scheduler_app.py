import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from io import BytesIO
import datetime

st.set_page_config(layout="wide", page_title="uKids Scheduler")
st.title("📅 uKids Volunteer Scheduler - September 2025")

# === Load uploaded files ===
avail_file = st.file_uploader("Upload Availability CSV", type=["csv"])
position_file = st.file_uploader("Upload Serving Positions CSV", type=["csv"])

if avail_file and position_file:
    avail_df = pd.read_csv(avail_file, encoding="ISO-8859-1")
    pos_df = pd.read_csv(position_file, encoding="ISO-8859-1")

    # === Extract relevant columns ===
    name_col = 'What is your name AND surname?'
    date_cols = [col for col in avail_df.columns if "September" in col]

    availability = {}
    for _, row in avail_df.iterrows():
        name = row[name_col].strip()
        availability[name] = {}
        for col in date_cols:
            date_str = col.split("available")[-1].strip()
            availability[name][date_str] = str(row[col]).strip().lower() == "yes"

    # Positions file
    pos_df = pos_df.rename(columns={pos_df.columns[0]: "Name"})
    pos_df = pos_df.dropna(subset=["Name"])
    pos_df["Name"] = pos_df["Name"].str.strip()

    roles = [col for col in pos_df.columns if not col.startswith("Unnamed") and col != "Name"]
    eligibility = {}
    for _, row in pos_df.iterrows():
        person = row["Name"]
        if person not in availability:
            continue
        eligibility[person] = {}
        for role in roles:
            try:
                pr = int(row[role])
            except:
                pr = 0
            if pr > 0:
                eligibility[person][role] = pr

    # Role capacities (from user instruction)
    capacities = {
        "Age 1 leader": 1,
        "Age 1 classroom": 5,
        "Age 1 nappies": 1,
        "Age 1 bags girls": 1,
        "Age 1 bags boys": 1,
        "Age 2 leader": 1,
        "Age 2 classroom": 4,
        "Age 2 nappies": 1,
        "Age 2 bags girls": 1,
        "Age 2 bags boys": 1,
        "Age 3 leader": 1,
        "Age 3 classroom": 4,
        "Age 3 bags": 1,
        "Age 4 leader": 1,
        "Age 4 classroom": 4,
        "Age 5 leader": 1,
        "Age 5 classroom": 3,
        "Age 6 leader": 1,
        "Age 6 classroom": 3,
        "Age 7 leader": 1,
        "Age 7 classroom": 2,
        "Age 8 leader": 1,
        "Age 8 classroom": 2,
        "Age 9 leader": 1,
        "Age 9 classroom": 1,
        "Age 10 leader": 1,
        "Age 10 classroom": 1,
        "Age 11 leader": 1,
        "Age 11 classroom": 1,
        "Special needs leader": 1,
        "Special needs classroom": 2
    }

    # Generate assignments
    from collections import defaultdict
    assignments = defaultdict(lambda: defaultdict(list))  # date -> role -> [names]
    count = defaultdict(int)  # total assigned per person

    for date in date_cols:
        date_str = date.split("available")[-1].strip()
        for role, needed in capacities.items():
            possible = [
                p for p in availability
                if availability[p].get(date_str, False)
                and role in eligibility.get(p, {})
                and eligibility[p][role] > 0
                and count[p] < 2
                and len(assignments[date_str][role]) < needed
            ]
            assigned = 0
            for p in sorted(possible, key=lambda x: count[x]):
                if assigned >= needed:
                    break
                assignments[date_str][role].append(p)
                count[p] += 1
                assigned += 1

    # === Output to styled Excel ===
    output = BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "September 2025"

    # Write header
    all_roles = list(capacities.keys())
    ws.cell(row=1, column=1, value="Role")
    for col, date in enumerate(date_cols, start=2):
        ws.cell(row=1, column=col, value=date.split("available")[-1].strip())

    # Write data
    for row_idx, role in enumerate(all_roles, start=2):
        ws.cell(row=row_idx, column=1, value=role)
        for col_idx, date in enumerate(date_cols, start=2):
            val = ", ".join(assignments[date.split("available")[-1].strip()][role])
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    # Autosize
    for col in ws.columns:
        max_len = max((len(str(cell.value)) if cell.value else 0) for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max(max_len + 2, 15)

    wb.save(output)
    st.download_button(
        label="📥 Download Schedule",
        data=output.getvalue(),
        file_name="ukids_september_2025_schedule.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    # Stats
    st.subheader("📊 Assignment Summary")
    stats = pd.DataFrame(sorted(count.items(), key=lambda x: x[0]), columns=["Volunteer", "# Assignments"])
    st.dataframe(stats, use_container_width=True)
